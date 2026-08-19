import pandas as pd
from pathlib import Path

input_path = Path("C:/Users/carlo/Downloads/Projetos V - Macro/Base de dados/Planilha de Séries Históricas - (Descontinuada).xlsx")
xls = pd.ExcelFile(input_path)

print("Abas encontradas:")
for i, s in enumerate(xls.sheet_names, 1):
    print(f"{i}. {s}")
# Inspecionar as 5 abas prioritárias: dimensões, primeiras linhas e possíveis cabeçalhos
sheets = [
    "NPL_Nova segmentação",
    "Carteira_Nova segm. e títulos",
    "BRGAAP - Operação e Risco",
    "BRGAAP - Vencimento e Risco",
    "PDD",
    "Créditos Renegociados"
]

for sheet in sheets:
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    print("\n" + "="*90)
    print(sheet, raw.shape)
    print(raw.iloc[:12, :12].to_string(index=False, header=False))
# Mapear datas/blocos e linhas das duas abas com muitos períodos
for sheet in ["BRGAAP - Operação e Risco", "BRGAAP - Vencimento e Risco"]:
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    print("\n", sheet)
    # datas presentes na primeira linha
    dates = []
    for c, v in enumerate(raw.iloc[0]):
        if pd.notna(v):
            try:
                d = pd.to_datetime(v)
                if 2000 <= d.year <= 2030:
                    dates.append((c, d.date()))
            except:
                pass
    print("Primeiras/últimas datas:", dates[:5], "...", dates[-5:])
    print("Número de blocos:", len(dates))
    print("Linhas:", raw.iloc[:,0].dropna().tolist())
for sheet in ["NPL_Nova segmentação", "Carteira_Nova segm. e títulos", "PDD"]:
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    print("\n", sheet, raw.shape)
    print("Datas iniciais:", [pd.to_datetime(v).date() for v in raw.iloc[1] if pd.notna(v) and not isinstance(v,str)][:5] if sheet != "PDD" else raw.iloc[1, :8].tolist())
    print("Últimas colunas com dados:")
    for c in range(raw.shape[1]-1, -1, -1):
        if raw.iloc[:,c].notna().any():
            print(c, raw.iloc[0,c], raw.iloc[1,c] if raw.shape[0]>1 else "")
            if c < raw.shape[1]-5: break
            # limited
            if c < raw.shape[1]-1: pass
    print("Linhas:")
    print(raw.iloc[:,0].dropna().tolist()[:25])
# Mostrar linhas-chave e alguns valores para entender unidades/percentuais
for sheet, row_names in {
    "NPL_Nova segmentação": [
        "Saldo NPL 90 dias - Pessoas Físicas - Brasil",
        "NPL 90 dias - Pessoas Físicas - Brasil",
        "Saldo NPL 15-90 dias - Pessoas Físicas - Brasil",
        "NPL 15-90 dias - Pessoas Físicas - Brasil"
    ],
    "Carteira_Nova segm. e títulos": [
        "Pessoas Físicas", "Cartão de Crédito", "Crédito Pessoal",
        "Crédito Consignado", "Veículos", "Crédito Imobiliário"
    ]
}.items():
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    print("\n", sheet)
    for name in row_names:
        matches = raw.index[raw.iloc[:,0].astype(str).str.strip() == name]
        for r in matches:
            print(name, "row", r, ":", raw.iloc[r,1:7].tolist())
# Criar a base integrada
import numpy as np

def read_simple_sheet(sheet, date_row=1):
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    # datas da linha indicada
    records = []
    for c in range(1, raw.shape[1]):
        v = raw.iloc[date_row, c]
        if pd.notna(v):
            try:
                d = pd.to_datetime(v)
            except:
                continue
            if 2000 <= d.year <= 2030:
                records.append((c, d))
    return raw, records

def extract_rows(sheet, row_names, date_row=1):
    raw, date_cols = read_simple_sheet(sheet, date_row)
    idx = {}
    labels = raw.iloc[:,0].astype(str).str.strip()
    for name in row_names:
        m = labels[labels == name]
        if len(m):
            idx[name] = m.index[0]
    data = []
    for c, d in date_cols:
        rec = {"date": d}
        for name, r in idx.items():
            rec[name] = pd.to_numeric(raw.iloc[r, c], errors="coerce")
        data.append(rec)
    return pd.DataFrame(data)

# NPL
npl_names = [
    "Saldo NPL 90 dias - Pessoas Físicas - Brasil",
    "NPL 90 dias - Pessoas Físicas - Brasil",
    "Saldo NPL 15-90 dias - Pessoas Físicas - Brasil",
    "NPL 15-90 dias - Pessoas Físicas - Brasil",
]
npl = extract_rows("NPL_Nova segmentação", npl_names)

# Carteira
cart_names = [
    "Pessoas Físicas", "Cartão de Crédito", "Crédito Pessoal",
    "Crédito Consignado", "Veículos", "Crédito Imobiliário",
    "Total Brasil com Garantias Financeiras Prestadas e Títulos Privados"
]
cart = extract_rows("Carteira_Nova segm. e títulos", cart_names)

print("NPL:", npl.shape, npl.date.min(), npl.date.max())
print("Carteira:", cart.shape, cart.date.min(), cart.date.max())
print(npl.tail(3).to_string(index=False))
def extract_risk_block(sheet):
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)

    labels = ["AA", "A", "B", "C", "D", "E", "F", "G", "H", "Total"]
    out = []

    for start in range(1, raw.shape[1], 10):

        # proteção para evitar blocos incompletos
        if start + len(labels) > raw.shape[1]:
            print(
                f"Bloco ignorado: start={start}, "
                f"última coluna necessária={start + len(labels) - 1}, "
                f"nº colunas={raw.shape[1]}"
            )
            continue

        d = raw.iloc[0, start]

        if pd.isna(d):
            continue

        try:
            d = pd.to_datetime(d)
        except (ValueError, TypeError):
            continue

        if not (2000 <= d.year <= 2030):
            continue

        vals = {}

        for idx, label in enumerate(labels):
            c = start + idx

            vals[label] = pd.to_numeric(
                raw.iloc[2, c],
                errors="coerce"
            )

        out.append({
            "date": d,
            **{f"risk_{k}": v for k, v in vals.items()}
        })

    return pd.DataFrame(out)


risk = extract_risk_block("BRGAAP - Operação e Risco")

risk = risk.rename(
    columns={"risk_Total": "risk_total_credit"}
)

risk["risk_high_share"] = (
    risk[["risk_D", "risk_E", "risk_F", "risk_G", "risk_H"]]
    .sum(axis=1)
    / risk["risk_total_credit"]
)

risk["risk_BH_share"] = (
    risk[
        ["risk_B", "risk_C", "risk_D", "risk_E",
         "risk_F", "risk_G", "risk_H"]
    ]
    .sum(axis=1)
    / risk["risk_total_credit"]
)
# Vencimento e risco: use "Operações em Curso Anormal" rows and total across risk levels
# ============================================================
# VENCIMENTO E RISCO
# ============================================================

import re
import unicodedata


def normalizar_texto(x):
    """
    Normaliza textos vindos do Excel:
    - converte para string
    - remove espaços invisíveis
    - remove espaços duplicados
    - remove acentos
    - converte para minúsculas
    """
    if pd.isna(x):
        return ""

    x = str(x)

    # espaço não separável do Excel
    x = x.replace("\xa0", " ")

    # espaços duplicados
    x = re.sub(r"\s+", " ", x).strip()

    # remover acentos
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))

    return x.lower()


def extract_vencimento(sheet):

    raw = pd.read_excel(
        input_path,
        sheet_name=sheet,
        header=None
    )

    # --------------------------------------------------------
    # Normalizar os nomes das linhas
    # --------------------------------------------------------

    labels_original = raw.iloc[:, 0].astype(str)

    labels = labels_original.apply(normalizar_texto)

    # --------------------------------------------------------
    # Linhas que queremos encontrar
    # --------------------------------------------------------

    wanted = {
        "abnormal_vencidas_total": "Parcelas Vencidas",
        "abnormal_01_14": "01 a 14",
        "abnormal_15_30": "15 a 30",
        "abnormal_31_60": "31 a 60",
        "abnormal_61_90": "61 a 90",
        "abnormal_91_180": "91 a 180",
        "abnormal_181_365": "181 a 365",
        "abnormal_above_365": "Acima de 365",
        "abnormal_subtotal": "Subtotal (a)",
        "normal_vencidas_14": "Parcelas Vencidas até 14 dias",
        "total_carteira": "Total da Carteira (a + b)",
        "provisao_existente": "Provisão Existente (3)"
    }

    rowmap = {}

    # --------------------------------------------------------
    # Procurar linhas
    # primeiro tenta igualdade exata;
    # se não encontrar, tenta busca parcial
    # --------------------------------------------------------

    for key, label in wanted.items():

        alvo = normalizar_texto(label)

        # 1. busca exata
        matches = list(
            labels[labels == alvo].index
        )

        # 2. busca parcial, caso o Excel tenha "dias", espaços etc.
        if not matches:

            matches = list(
                labels[
                    labels.str.contains(
                        re.escape(alvo),
                        na=False
                    )
                ].index
            )

        if matches:

            if key == "normal_vencidas_14":
                # usar última ocorrência, caso haja repetição
                rowmap[key] = matches[-1]

            else:
                rowmap[key] = matches[0]

        else:

            print(
                f"AVISO: linha não encontrada -> "
                f"{key}: '{label}'"
            )

    # --------------------------------------------------------
    # Mostrar o que foi encontrado
    # --------------------------------------------------------

    print("\nLinhas encontradas em Vencimento e Risco:")

    for key, r in rowmap.items():

        print(
            f"{key:30s} -> "
            f"linha {r}: "
            f"{labels_original.iloc[r]}"
        )

    # --------------------------------------------------------
    # Encontrar períodos
    # --------------------------------------------------------

    out = []

    labels_risco = [
        "AA", "A", "B", "C", "D",
        "E", "F", "G", "H", "Total"
    ]

    for start in range(1, raw.shape[1], 10):

        # Proteção contra último bloco incompleto
        if start + 9 >= raw.shape[1]:
            continue

        d = raw.iloc[0, start]

        if pd.isna(d):
            continue

        try:
            d = pd.to_datetime(d)

        except (ValueError, TypeError):
            continue

        if not (2000 <= d.year <= 2030):
            continue

        rec = {
            "date": d
        }

        # Total é a última coluna do bloco
        total_col = start + 9

        for key, r in rowmap.items():

            rec[key] = pd.to_numeric(
                raw.iloc[r, total_col],
                errors="coerce"
            )

        out.append(rec)

    venc = pd.DataFrame(out)

    # --------------------------------------------------------
    # Garantir que todas as colunas necessárias existam
    # --------------------------------------------------------

    required = [
        "abnormal_vencidas_total",
        "abnormal_01_14",
        "abnormal_15_30",
        "abnormal_31_60",
        "abnormal_61_90",
        "abnormal_91_180",
        "abnormal_181_365",
        "abnormal_above_365",
        "abnormal_subtotal",
        "normal_vencidas_14",
        "total_carteira",
        "provisao_existente"
    ]

    for col in required:

        if col not in venc.columns:
            venc[col] = pd.NA

    return venc


# ============================================================
# EXTRAÇÃO
# ============================================================

venc = extract_vencimento(
    "BRGAAP - Vencimento e Risco"
)


# ============================================================
# VARIÁVEIS DERIVADAS
# ============================================================

venc["early_delinquency_1_30"] = (
    pd.to_numeric(
        venc["abnormal_01_14"],
        errors="coerce"
    )
    +
    pd.to_numeric(
        venc["abnormal_15_30"],
        errors="coerce"
    )
)


venc["severe_delinquency_91_plus"] = (
    venc[
        [
            "abnormal_91_180",
            "abnormal_181_365",
            "abnormal_above_365"
        ]
    ]
    .apply(pd.to_numeric, errors="coerce")
    .sum(axis=1, min_count=1)
)


venc["overdue_share"] = (
    pd.to_numeric(
        venc["abnormal_vencidas_total"],
        errors="coerce"
    )
    /
    pd.to_numeric(
        venc["total_carteira"],
        errors="coerce"
    )
)


venc["severe_overdue_share"] = (
    venc["severe_delinquency_91_plus"]
    /
    pd.to_numeric(
        venc["total_carteira"],
        errors="coerce"
    )
)


print("\nDimensão da base de vencimentos:")
print(venc.shape)

print("\nColunas:")
print(venc.columns.tolist())

print("\nÚltimas observações:")
print(venc.tail().to_string(index=False))

venc = extract_vencimento("BRGAAP - Vencimento e Risco")
venc["early_delinquency_1_30"] = venc["abnormal_01_14"] + venc["abnormal_15_30"]
venc["severe_delinquency_91_plus"] = venc[["abnormal_91_180","abnormal_181_365","abnormal_above_365"]].sum(axis=1)
venc["overdue_share"] = venc["abnormal_vencidas_total"] / venc["total_carteira"]
venc["severe_overdue_share"] = venc["severe_delinquency_91_plus"] / venc["total_carteira"]

print("risk", risk.shape, risk.date.min(), risk.date.max())
print("venc", venc.shape, venc.date.min(), venc.date.max())
print(venc.tail(2).to_string(index=False))
raw = pd.read_excel(input_path, sheet_name="BRGAAP - Operação e Risco", header=None)
print(list(enumerate(raw.iloc[0, 0:13].tolist())))
print(list(enumerate(raw.iloc[1, 0:13].tolist())))
def extract_risk_block(sheet):
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    labels = ["AA","A","B","C","D","E","F","G","H","Total"]
    out = []
    for start in range(1, raw.shape[1], 10):
        d = raw.iloc[0, start]
        if pd.isna(d):
            continue
        try:
            d = pd.to_datetime(d)
        except:
            continue
        if not (2000 <= d.year <= 2030):
            continue
        vals = {}
        for idx, label in enumerate(labels):
            c = start + idx
            vals[label] = pd.to_numeric(raw.iloc[2, c], errors="coerce")
        out.append({"date":d, **{f"risk_{k}":v for k,v in vals.items()}})
    return pd.DataFrame(out)

risk = extract_risk_block("BRGAAP - Operação e Risco")
risk = risk.rename(columns={"risk_Total":"risk_total_credit"})
risk["risk_high_share"] = risk[["risk_D","risk_E","risk_F","risk_G","risk_H"]].sum(axis=1) / risk["risk_total_credit"]
risk["risk_BH_share"] = risk[["risk_B","risk_C","risk_D","risk_E","risk_F","risk_G","risk_H"]].sum(axis=1) / risk["risk_total_credit"]

def extract_vencimento(sheet):
    raw = pd.read_excel(input_path, sheet_name=sheet, header=None)
    labels = raw.iloc[:,0].astype(str).str.strip()
    wanted = {
        "abnormal_vencidas_total":"Parcelas Vencidas",
        "abnormal_01_14":"01 a 14",
        "abnormal_15_30":"15 a 30",
        "abnormal_31_60":"31 a 60",
        "abnormal_61_90":"61 a 90",
        "abnormal_91_180":"91 a 180",
        "abnormal_181_365":"181 a 365",
        "abnormal_above_365":"Acima de 365",
        "abnormal_subtotal":"Subtotal (a)",
        "normal_vencidas_14":"Parcelas Vencidas até 14 dias",
        "total_carteira":"Total da Carteira (a + b)",
        "provisao_existente":"Provisão Existente (3)"
    }
    rowmap={}
    for key,label in wanted.items():
        matches = list(labels[labels == label].index)
        if matches:
            rowmap[key] = matches[-1] if key=="normal_vencidas_14" else matches[0]
    out=[]
    for start in range(1, raw.shape[1], 10):
        d=raw.iloc[0,start]
        if pd.isna(d): continue
        try: d=pd.to_datetime(d)
        except: continue
        if not (2000 <= d.year <= 2030): continue
        rec={"date":d}
        for key,r in rowmap.items():
            rec[key]=pd.to_numeric(raw.iloc[r,start+9], errors="coerce")
        out.append(rec)
    return pd.DataFrame(out)

venc = extract_vencimento("BRGAAP - Vencimento e Risco")
venc["early_delinquency_1_30"] = venc["abnormal_01_14"] + venc["abnormal_15_30"]
venc["severe_delinquency_91_plus"] = venc[["abnormal_91_180","abnormal_181_365","abnormal_above_365"]].sum(axis=1)
venc["overdue_share"] = venc["abnormal_vencidas_total"] / venc["total_carteira"]
venc["severe_overdue_share"] = venc["severe_delinquency_91_plus"] / venc["total_carteira"]

print(risk.tail(2).to_string(index=False))
print(venc.tail(2).to_string(index=False))
# PDD e créditos renegociados
def parse_quarter_label(x):
    if pd.isna(x): return pd.NaT
    s = str(x).strip().upper()
    import re
    m = re.match(r"([1-4])T(\d{2})$", s)
    if not m: return pd.NaT
    q, yy = int(m.group(1)), int(m.group(2))
    year = 2000 + yy
    month = {1:3, 2:6, 3:9, 4:12}[q]
    return pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(1)

def extract_pdd():
    raw = pd.read_excel(input_path, sheet_name="PDD", header=None)
    labels = raw.iloc[:,0].astype(str).str.strip()
    wanted = {
        "pdd_constituicao_liquida":"Constituição Líquida no Período",
        "pdd_writeoffs":"Write-Offs",
        "pdd_saldo_final":"Saldo Final",
        "pdd_minima":"Mínima",
        "pdd_garantias":"Provisão Para Garantias Financ. Prestadas",
        "pdd_complementar":"Provisão Complementar",
    }
    rowmap={}
    for key,label in wanted.items():
        m=labels[labels==label]
        if len(m): rowmap[key]=m.index[0]
    out=[]
    for c in range(1, raw.shape[1]):
        d=parse_quarter_label(raw.iloc[1,c])
        if pd.isna(d): continue
        rec={"date":d}
        for key,r in rowmap.items():
            rec[key]=pd.to_numeric(raw.iloc[r,c], errors="coerce")
        out.append(rec)
    return pd.DataFrame(out)

pdd = extract_pdd()

def extract_reneg():
    raw = pd.read_excel(input_path, sheet_name="Créditos Renegociados", header=None)
    labels=raw.iloc[:,0].astype(str).str.strip()
    wanted={
        "reneg_total":"Créditos Renegociados Totais",
        "reneg_vencidos_ate30":"Créditos Renegociados vencidos até 30 dias *",
        "reneg_vencidos_acima30":"Créditos Renegociados vencidos acima de 30 dias *",
        "reneg_total_brasil":"Créditos Renegociados Total - Brasil",
    }
    rowmap={}
    for key,label in wanted.items():
        m=labels[labels==label]
        if len(m): rowmap[key]=m.index[0]
    out=[]
    for c in range(1,raw.shape[1]):
        v=raw.iloc[1,c]
        try: d=pd.to_datetime(v)
        except: continue
        if pd.isna(d): continue
        rec={"date":d}
        for key,r in rowmap.items():
            rec[key]=pd.to_numeric(raw.iloc[r,c],errors="coerce")
        out.append(rec)
    return pd.DataFrame(out)

reneg=extract_reneg()

print("PDD:", pdd.shape, pdd.date.min(), pdd.date.max())
print(pdd.tail(2).to_string(index=False))
print("Reneg:", reneg.shape, reneg.date.min(), reneg.date.max())
# Construção da base final
df = (
    npl.merge(cart, on="date", how="inner")
       .merge(risk[["date","risk_total_credit","risk_high_share","risk_BH_share"]], on="date", how="inner")
       .merge(venc[["date","early_delinquency_1_30","severe_delinquency_91_plus",
                    "overdue_share","severe_overdue_share","provisao_existente"]], on="date", how="inner")
       .merge(pdd, on="date", how="inner")
       .merge(reneg, on="date", how="inner")
)

# Variáveis derivadas
df["npl90_pf_rate_calc"] = (
    df["Saldo NPL 90 dias - Pessoas Físicas - Brasil"] / df["Pessoas Físicas"]
)
df["npl90_pf_rate"] = df["NPL 90 dias - Pessoas Físicas - Brasil"]
df["npl15_90_pf_rate"] = df["NPL 15-90 dias - Pessoas Físicas - Brasil"]

# Composição da carteira PF
for col, new in [
    ("Cartão de Crédito","share_cartao_pf"),
    ("Crédito Pessoal","share_credito_pessoal_pf"),
    ("Crédito Consignado","share_consignado_pf"),
    ("Veículos","share_veiculos_pf"),
    ("Crédito Imobiliário","share_imobiliario_pf")
]:
    df[new] = df[col] / df["Pessoas Físicas"]

# Indicadores consolidados (não específicos de PF)
df["pdd_to_credit_rate"] = abs(df["pdd_saldo_final"]) / df["risk_total_credit"]
df["reneg_to_credit_rate"] = df["reneg_total_brasil"] / df["risk_total_credit"]

# Variações trimestrais de algumas variáveis
for col in ["npl90_pf_rate", "npl15_90_pf_rate", "risk_high_share",
            "overdue_share", "severe_overdue_share", "pdd_to_credit_rate",
            "reneg_to_credit_rate"]:
    df[f"delta_{col}"] = df[col].diff()

# Identificação temporal
df["year"] = df["date"].dt.year
df["quarter"] = df["date"].dt.quarter
df["period"] = df["date"].dt.to_period("Q").astype(str)

# Seleção e nomes amigáveis para a base analítica
final_cols = [
    "date","period","year","quarter",
    "Pessoas Físicas",
    "Cartão de Crédito","Crédito Pessoal","Crédito Consignado","Veículos","Crédito Imobiliário",
    "Saldo NPL 90 dias - Pessoas Físicas - Brasil","npl90_pf_rate","npl90_pf_rate_calc",
    "Saldo NPL 15-90 dias - Pessoas Físicas - Brasil","npl15_90_pf_rate",
    "share_cartao_pf","share_credito_pessoal_pf","share_consignado_pf",
    "share_veiculos_pf","share_imobiliario_pf",
    "risk_total_credit","risk_high_share","risk_BH_share",
    "early_delinquency_1_30","severe_delinquency_91_plus",
    "overdue_share","severe_overdue_share",
    "pdd_saldo_final","pdd_constituicao_liquida","pdd_writeoffs",
    "pdd_to_credit_rate",
    "reneg_total_brasil","reneg_vencidos_ate30","reneg_vencidos_acima30",
    "reneg_to_credit_rate"
]
final = df[final_cols].sort_values("date").reset_index(drop=True)

# Renomear colunas para nomes adequados ao uso em Python/SQL/R
rename = {
    "Pessoas Físicas":"credit_pf",
    "Cartão de Crédito":"credit_card_pf",
    "Crédito Pessoal":"personal_credit_pf",
    "Crédito Consignado":"payroll_credit_pf",
    "Veículos":"vehicle_credit_pf",
    "Crédito Imobiliário":"mortgage_credit_pf",
    "Saldo NPL 90 dias - Pessoas Físicas - Brasil":"npl90_balance_pf",
    "Saldo NPL 15-90 dias - Pessoas Físicas - Brasil":"npl15_90_balance_pf",
    "Saldo NPL 90 dias - Pessoas Físicas - Brasil":"npl90_balance_pf",
    "pdd_saldo_final":"pdd_balance_total",
    "pdd_constituicao_liquida":"pdd_net_formation",
    "pdd_writeoffs":"pdd_writeoffs",
    "reneg_total_brasil":"renegotiated_credit_brazil",
    "reneg_vencidos_ate30":"renegotiated_overdue_upto30",
    "reneg_vencidos_acima30":"renegotiated_overdue_above30",
}
final = final.rename(columns=rename)

# Percentuais em formato decimal; deixar os saldos em R$ milhões conforme as abas de origem.
# Validação: taxa calculada a partir do saldo deve coincidir com a divulgada.
final["npl90_rate_gap"] = final["npl90_pf_rate"] - final["npl90_pf_rate_calc"]

# Metadados
metadata = pd.DataFrame({
    "item": [
        "Período da base final",
        "Frequência",
        "Unidade dos saldos",
        "NPL PF",
        "Indicadores de risco/vencimento",
        "PDD e renegociação",
        "Observação metodológica"
    ],
    "valor": [
        f"{final.date.min():%Y-%m-%d} a {final.date.max():%Y-%m-%d}",
        "Trimestral",
        "R$ milhões, conforme divulgado pelo Itaú",
        "Específico para Pessoas Físicas - Brasil",
        "Consolidados da carteira com característica de concessão de crédito; não são específicos de PF",
        "Consolidados; usados como indicadores complementares de qualidade da carteira",
        "A base é um painel temporal agregado. Ela não identifica individualmente quem aposta em bets."
    ]
})

# ============================================================
# SALVAR CSV E EXCEL
# ============================================================

out_dir = Path(
    "C:/Users/carlo/Downloads/Projetos V - Macro/Base de dados/base_consolidada_output"
)

# garante que a pasta existe
out_dir.mkdir(parents=True, exist_ok=True)

csv_path = out_dir / "itau_base_inadimplencia_pf.csv"
xlsx_path = out_dir / "itau_base_inadimplencia_pf.xlsx"

# salvar CSV
final.to_csv(
    csv_path,
    index=False,
    encoding="utf-8-sig"
)

# salvar Excel
with pd.ExcelWriter(
    xlsx_path,
    engine="openpyxl"
) as writer:

    final.to_excel(
        writer,
        sheet_name="base_analitica",
        index=False
    )

    metadata.to_excel(
        writer,
        sheet_name="metadados",
        index=False
    )

print("\nArquivos criados com sucesso:")
print(f"CSV:   {csv_path}")
print(f"Excel: {xlsx_path}")

# Formatação básica da planilha
from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
ws = wb["base_analitica"]
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 28)
wb.save(xlsx_path)

print(f"Base criada com {len(final)} observações e {len(final.columns)} variáveis.")
print(f"CSV: {csv_path}")
print(f"Excel: {xlsx_path}")
print("\nPrimeiras linhas:")
print(final.head(3).to_string(index=False))
print("\nÚltimas linhas:")
print(final.tail(3).to_string(index=False))
print("\nMaior diferença entre NPL divulgado e calculado:", final["npl90_rate_gap"].abs().max())
# Adicionar um dicionário de dados ao Excel para deixar a base pronta para o grupo
dictionary = pd.DataFrame([
    ("date","Data","Data de referência trimestral"),
    ("period","Período","Período no formato AAAAQT"),
    ("credit_pf","Carteira PF","Carteira de crédito de Pessoas Físicas; R$ milhões"),
    ("credit_card_pf","Cartão PF","Carteira de cartão de crédito; R$ milhões"),
    ("personal_credit_pf","Crédito pessoal PF","Carteira de crédito pessoal; R$ milhões"),
    ("payroll_credit_pf","Consignado PF","Carteira de crédito consignado; R$ milhões"),
    ("vehicle_credit_pf","Veículos PF","Carteira de financiamento de veículos; R$ milhões"),
    ("mortgage_credit_pf","Imobiliário PF","Carteira de crédito imobiliário; R$ milhões"),
    ("npl90_balance_pf","Saldo NPL 90 PF","Saldo de NPL 90 dias de PF no Brasil; R$ milhões"),
    ("npl90_pf_rate","Taxa NPL 90 PF","Taxa NPL 90 dias de PF divulgada pelo Itaú"),
    ("npl90_pf_rate_calc","Taxa NPL 90 PF calculada","Saldo NPL PF / carteira PF; serve como checagem, não substitui a taxa divulgada"),
    ("npl15_90_balance_pf","Saldo NPL 15-90 PF","Saldo NPL 15-90 dias de PF; R$ milhões"),
    ("npl15_90_pf_rate","Taxa NPL 15-90 PF","Taxa NPL 15-90 dias de PF divulgada pelo Itaú"),
    ("share_cartao_pf","Share cartão PF","Cartão / carteira PF"),
    ("share_credito_pessoal_pf","Share crédito pessoal PF","Crédito pessoal / carteira PF"),
    ("share_consignado_pf","Share consignado PF","Consignado / carteira PF"),
    ("share_veiculos_pf","Share veículos PF","Veículos / carteira PF"),
    ("share_imobiliario_pf","Share imobiliário PF","Imobiliário / carteira PF"),
    ("risk_total_credit","Carteira de crédito para risco","Operações de crédito usadas na aba Operação e Risco; R$ milhões"),
    ("risk_high_share","Share risco D-H","(D+E+F+G+H) / operações de crédito; indicador consolidado"),
    ("risk_BH_share","Share risco B-H","(B+C+D+E+F+G+H) / operações de crédito; indicador consolidado"),
    ("early_delinquency_1_30","Atrasos 1-30 dias","Atrasos de 1 a 30 dias na carteira em curso anormal; R$ milhões"),
    ("severe_delinquency_91_plus","Atrasos >90 dias","Atrasos de 91 dias ou mais em curso anormal; R$ milhões"),
    ("overdue_share","Share vencido","Parcelas vencidas / total da carteira; indicador consolidado"),
    ("severe_overdue_share","Share >90 dias","Atrasos >90 dias / total da carteira; indicador consolidado"),
    ("pdd_balance_total","Saldo PDD","Saldo final da provisão; R$ milhões; valor negativo conforme divulgação"),
    ("pdd_net_formation","Constituição líquida PDD","Constituição líquida no período; R$ milhões"),
    ("pdd_writeoffs","Write-offs","Baixas por write-off; R$ milhões"),
    ("pdd_to_credit_rate","PDD / crédito","|Saldo PDD| / operações de crédito; indicador consolidado"),
    ("renegotiated_credit_brazil","Crédito renegociado Brasil","Créditos renegociados no Brasil; R$ milhões"),
    ("renegotiated_overdue_upto30","Renegociado vencido até 30","Créditos renegociados vencidos até 30 dias; R$ milhões"),
    ("renegotiated_overdue_above30","Renegociado vencido >30","Créditos renegociados vencidos acima de 30 dias; R$ milhões"),
    ("reneg_to_credit_rate","Renegociado / crédito","Créditos renegociados Brasil / operações de crédito; indicador consolidado"),
    ("npl90_rate_gap","Diferença NPL","Taxa NPL divulgada menos saldo NPL / carteira PF; diagnóstico de denominador/escopo")
], columns=["variavel","descricao","definicao"])

with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    dictionary.to_excel(writer, sheet_name="dicionario", index=False)

wb = load_workbook(xlsx_path)
ws = wb["dicionario"]
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 65)
wb.save(xlsx_path)

print("Dicionário de dados adicionado ao Excel.")
